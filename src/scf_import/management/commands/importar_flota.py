"""
Django Management Command: importar_flota
========================================
Permite importar y sincronizar periódicamente la flota vehicular en el SCF
a partir de un archivo CSV o Excel normalizado (flota_normalizada.YYYY-MM-DD.xlsx / .csv).

Características:
- Idempotente: Realiza Upsert (update_or_create) por clave natural (VIN o SAP).
- Cero Dependencias Pesadas: Usa exclusivamente la biblioteca estándar (csv) y openpyxl.
- Preserva Historial: Mantiene la Primary Key (ID) del vehículo en la base de datos,
  garantizando que no se pierdan órdenes de trabajo ni mantenimientos asociados.
- Resuelve Catálogos: Realiza get_or_create dinámico para marcas, modelos, estados,
  gerencias y centros de servicio según los modelos canónicos de SCF.
- Soporta --dry-run para simular sin persistir cambios en la base de datos.
"""
import csv
from pathlib import Path
from typing import Any

try:
    from django.core.management.base import BaseCommand, CommandError
    from django.db import transaction
    DJANGO_AVAILABLE = True
except ImportError:
    DJANGO_AVAILABLE = False
    BaseCommand = object  # type: ignore
    CommandError = Exception  # type: ignore


def _load_data_file(file_path: Path) -> list[dict[str, str]]:
    ext = file_path.suffix.lower()
    if ext == ".csv":
        with file_path.open("r", encoding="utf-8-sig") as f:
            sample = f.read(4096)
            f.seek(0)
            delim = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.DictReader(f, delimiter=delim)
            return [dict(row) for row in reader]
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl es requerido para importar archivos .xlsx") from exc

        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for r in rows[1:]:
            row_dict = {}
            for col_idx, h in enumerate(headers):
                val = r[col_idx] if col_idx < len(r) else ""
                row_dict[h] = str(val).strip() if val is not None else ""
            result.append(row_dict)
        wb.close()
        return result
    else:
        raise ValueError(f"Formato no soportado ({ext}). Use .xlsx o .csv")


MAX_VIN_LENGTH = 17


class Command(BaseCommand):  # type: ignore[misc]
    help = "Importa o actualiza el inventario de flota en el SCF desde un CSV o Excel normalizado"

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument(
            "file_path",
            type=str,
            help="Ruta al archivo normalizado (flota_normalizada.YYYY-MM-DD.xlsx o .csv)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula la importación sin guardar cambios en la base de datos",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=500,
            help="Tamaño de lote para transacciones de base de datos (por defecto: 500)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        if not DJANGO_AVAILABLE:
            raise RuntimeError("Django no está instalado en este entorno.")

        file_path = Path(options["file_path"])
        dry_run = options["dry_run"]
        batch_size = options["batch_size"]

        if not file_path.exists():
            raise CommandError(f"El archivo especificado no existe: {file_path}")

        self.stdout.write(self.style.MIGRATE_HEADING(f"Iniciando importación desde: {file_path}"))
        if dry_run:
            self.stdout.write(self.style.WARNING("Modo --dry-run activado: No se guardarán cambios en la base de datos."))

        try:
            records = _load_data_file(file_path)
        except Exception as e:
            raise CommandError(f"Error al leer archivo: {e}")

        total_rows = len(records)
        self.stdout.write(f"Total de registros a procesar: {total_rows:,}")

        # 2. Importar modelos de Django dinámicamente según el esquema canónico de SCF
        try:
            from django.apps import apps
            Marca = apps.get_model("catalogos", "Marca")
            Modelo = apps.get_model("catalogos", "Modelo")
            Color = apps.get_model("catalogos", "Color")
            ClaseVehiculo = apps.get_model("catalogos", "ClaseVehiculo")
            TipoVehiculo = apps.get_model("catalogos", "TipoVehiculo")
            TipoCombustible = apps.get_model("catalogos", "TipoCombustible")
            EstatusVehiculo = apps.get_model("catalogos", "EstatusVehiculo")
            TipoUso = apps.get_model("catalogos", "TipoUso")
            ColorPlaca = apps.get_model("catalogos", "ColorPlaca")
            Estado = apps.get_model("organizacion", "Estado")
            Gerencia = apps.get_model("organizacion", "Gerencia")
            CentroDeServicio = apps.get_model("organizacion", "CentroDeServicio")
            Vehiculo = apps.get_model("vehiculos", "Vehiculo")
        except LookupError as e:
            raise CommandError(f"Error al cargar modelos de Django del SCF: {e}")

        # 3. Procesar registros
        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        # Cachés locales en memoria para optimizar get_or_create
        cache_marca: dict[str, Any] = {}
        cache_modelo: dict[tuple[int, str], Any] = {}
        cache_color: dict[str, Any] = {}
        cache_clase: dict[str, Any] = {}
        cache_categoria: dict[str, Any] = {}
        cache_combustible: dict[str, Any] = {}
        cache_estatus: dict[str, Any] = {}
        cache_uso: dict[str, Any] = {}
        cache_color_placa: dict[str, Any] = {}
        cache_estado: dict[str, Any] = {}
        cache_gerencia: dict[str, Any] = {}
        cache_centro_servicio: dict[str, Any] = {}

        try:
            with transaction.atomic():
                for idx, row in enumerate(records):
                    row_num = idx + 2
                    vin = str(row.get("vin", "")).strip().upper()
                    sap = str(row.get("numero_economico", "")).strip()
                    placa = str(row.get("placa", "")).strip().upper()

                    # Verificar identificadores requeridos
                    if not vin or len(vin) > MAX_VIN_LENGTH or not sap:
                        skipped_count += 1
                        continue

                    try:
                        # Resolver Catálogos
                        marca_name = str(row.get("marca", "")).strip().title()
                        if not marca_name:
                            skipped_count += 1
                            continue
                        if marca_name not in cache_marca:
                            marca_obj, _ = Marca.objects.get_or_create(nombre=marca_name)
                            cache_marca[marca_name] = marca_obj
                        else:
                            marca_obj = cache_marca[marca_name]

                        modelo_name = str(row.get("modelo", "")).strip()
                        if not modelo_name:
                            skipped_count += 1
                            continue
                        key_m = (marca_obj.pk, modelo_name)
                        if key_m not in cache_modelo:
                            modelo_obj, _ = Modelo.objects.get_or_create(
                                marca=marca_obj,
                                nombre=modelo_name,
                            )
                            cache_modelo[key_m] = modelo_obj
                        else:
                            modelo_obj = cache_modelo[key_m]

                        color_name = str(row.get("color", "")).strip().title()
                        color_obj = None
                        if color_name:
                            if color_name not in cache_color:
                                color_obj, _ = Color.objects.get_or_create(nombre=color_name)
                                cache_color[color_name] = color_obj
                            else:
                                color_obj = cache_color[color_name]

                        cp_name = str(row.get("color_placa", "")).strip().title()
                        color_placa_obj = None
                        if cp_name:
                            if cp_name not in cache_color_placa:
                                color_placa_obj, _ = ColorPlaca.objects.get_or_create(nombre=cp_name)
                                cache_color_placa[cp_name] = color_placa_obj
                            else:
                                color_placa_obj = cache_color_placa[cp_name]

                        clase_name = str(row.get("clase", "")).strip().title() or "Liviano"
                        if clase_name not in cache_clase:
                            clase_obj, _ = ClaseVehiculo.objects.get_or_create(nombre=clase_name)
                            cache_clase[clase_name] = clase_obj
                        else:
                            clase_obj = cache_clase[clase_name]

                        cat_name = str(row.get("categoria", "")).strip().title() or "Sedan"
                        if cat_name not in cache_categoria:
                            cat_obj, _ = TipoVehiculo.objects.get_or_create(nombre=cat_name)
                            cache_categoria[cat_name] = cat_obj
                        else:
                            cat_obj = cache_categoria[cat_name]

                        comb_name = str(row.get("tipo_combustible", "")).strip().title() or "Gasolina"
                        if comb_name not in cache_combustible:
                            comb_obj, _ = TipoCombustible.objects.get_or_create(nombre=comb_name)
                            cache_combustible[comb_name] = comb_obj
                        else:
                            comb_obj = cache_combustible[comb_name]

                        estatus_name = str(row.get("estatus", "")).strip().title() or "Operativo"
                        if estatus_name not in cache_estatus:
                            estatus_obj, _ = EstatusVehiculo.objects.get_or_create(nombre=estatus_name)
                            cache_estatus[estatus_name] = estatus_obj
                        else:
                            estatus_obj = cache_estatus[estatus_name]

                        uso_name = str(row.get("tipo_uso", "")).strip().title()
                        uso_obj = None
                        if uso_name:
                            if uso_name not in cache_uso:
                                uso_obj, _ = TipoUso.objects.get_or_create(nombre=uso_name)
                                cache_uso[uso_name] = uso_obj
                            else:
                                uso_obj = cache_uso[uso_name]

                        # Resolver Organización
                        estado_name = str(row.get("estado", "")).strip().title()
                        if not estado_name:
                            skipped_count += 1
                            continue
                        if estado_name not in cache_estado:
                            estado_obj, _ = Estado.objects.get_or_create(nombre=estado_name)
                            cache_estado[estado_name] = estado_obj
                        else:
                            estado_obj = cache_estado[estado_name]

                        gerencia_name = str(row.get("gerencia", "")).strip().title()
                        if not gerencia_name:
                            skipped_count += 1
                            continue
                        if gerencia_name not in cache_gerencia:
                            gerencia_obj, _ = Gerencia.objects.get_or_create(nombre=gerencia_name)
                            cache_gerencia[gerencia_name] = gerencia_obj
                        else:
                            gerencia_obj = cache_gerencia[gerencia_name]

                        uu_name = str(row.get("unidad_usuaria", "")).strip().title()
                        uu_obj = None
                        if uu_name:
                            if uu_name not in cache_gerencia:
                                uu_obj, _ = Gerencia.objects.get_or_create(nombre=uu_name)
                                cache_gerencia[uu_name] = uu_obj
                            else:
                                uu_obj = cache_gerencia[uu_name]

                        empl_name = str(row.get("emplazamiento", "")).strip().title()
                        if not empl_name:
                            skipped_count += 1
                            continue
                        if empl_name not in cache_centro_servicio:
                            empl_obj, _ = CentroDeServicio.objects.get_or_create(
                                nombre=empl_name,
                                defaults={"estado": estado_obj},
                            )
                            cache_centro_servicio[empl_name] = empl_obj
                        else:
                            empl_obj = cache_centro_servicio[empl_name]

                        # Año
                        try:
                            anio_raw = str(row.get("anio", "")).strip()
                            anio_val = int(anio_raw) if anio_raw.isdigit() else 2000
                        except (ValueError, TypeError):
                            anio_val = 2000

                        # Número de Unidad (Unique nullable)
                        unidad_str = str(row.get("numero_unidad", "")).strip() or None
                        if unidad_str and Vehiculo.objects.filter(numero_unidad=unidad_str).exclude(vin=vin).exists():
                            unidad_str = None

                        placa_str = placa or None

                        # Buscar vehículo existente por clave natural
                        vehiculo = Vehiculo.objects.filter(vin=vin).first()
                        if not vehiculo and sap:
                            vehiculo = Vehiculo.objects.filter(numero_economico=sap).first()

                        # Preparar campos a persistir compatibles con SCF
                        vehicle_data = {
                            "numero_economico": sap,
                            "vin": vin,
                            "placa": placa_str,
                            "color_placa": color_placa_obj,
                            "placa_intt": str(row.get("placa_intt", "")).strip(),
                            "serial_motor": str(row.get("serial_motor", "")).strip(),
                            "numero_unidad": unidad_str,
                            "anio": anio_val,
                            "marca": marca_obj,
                            "modelo": modelo_obj,
                            "color": color_obj,
                            "clase": clase_obj,
                            "categoria": cat_obj,
                            "tipo_combustible": comb_obj,
                            "estatus": estatus_obj,
                            "tipo_uso": uso_obj,
                            "estado": estado_obj,
                            "gerencia": gerencia_obj,
                            "unidad_usuaria": uu_obj,
                            "emplazamiento": empl_obj,
                            "estatus_activo": True,
                        }

                        if vehiculo:
                            for key, val in vehicle_data.items():
                                setattr(vehiculo, key, val)
                            if not dry_run:
                                vehiculo.save()
                            updated_count += 1
                        else:
                            if not dry_run:
                                Vehiculo.objects.create(**vehicle_data)
                            created_count += 1

                    except Exception as err:
                        error_count += 1
                        self.stdout.write(self.style.ERROR(f"Fila {row_num}: Error procesando vehículo - {err}"))

                if dry_run:
                    transaction.set_rollback(True)

        except Exception as e:
            raise CommandError(f"Fallo crítico en la transacción de base de datos: {e}")

        self.stdout.write("\n" + "=" * 50)
        self.stdout.write(self.style.SUCCESS("  RESUMEN DE SINCRONIZACIÓN DE FLOTA"))
        self.stdout.write("=" * 50)
        self.stdout.write(f"Total filas procesadas   : {total_rows:,}")
        self.stdout.write(f"Vehículos actualizados   : {updated_count:,}")
        self.stdout.write(f"Vehículos nuevos creados : {created_count:,}")
        self.stdout.write(f"Filas omitidas (sin id)  : {skipped_count:,}")
        self.stdout.write(f"Errores en filas         : {error_count:,}")
        self.stdout.write("=" * 50 + "\n")
